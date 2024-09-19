"""
Utility functions for saving and loading newsletter URLs to and from an Excel file.
"""
import os
from typing import List

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Save newsletter URLs to an Excel file
def save_newsletters_to_excel(newsletter_urls: List[str], file_name="subscribed_newsletters.xlsx") -> None:
    """
    Save the list of newsletter URLs to an Excel file.

    :param newsletter_urls: List of newsletter URLs to save
    :param file_name: Name of the Excel file
    """
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet: Worksheet | None = workbook.active
        if sheet:
            # Add a header row
            cell = sheet.cell(row=1, column=1, value="Newsletter URLs")
            cell.font = cell.font.copy(bold=True)
        workbook.save(filename=file_name)

    workbook: Workbook = openpyxl.load_workbook(filename=file_name)
    sheet: Worksheet | None = workbook.active

    for url in newsletter_urls:
        if sheet:
            sheet.append([url])

    workbook.save(filename=file_name)
    print(f"Saved subscribed newsletters to {file_name}")


# Load Newsletter URLs from the Excel file
def load_newsletters_from_excel(file_name: str = "subscribed_newsletters.xlsx") -> list[str]:
    """
    Load newsletter URLs from an Excel file.

    :param file_name: Name of the Excel file
    :return: List of newsletter URLs
    """
    if not os.path.exists(file_name):
        print(f"File {file_name} not found.")
        return []

    workbook: Workbook = openpyxl.load_workbook(filename=file_name)
    sheet: Worksheet | None = workbook.active

    urls = []
    if sheet:
        for row in sheet.iter_rows(values_only=True):
            urls.append(row[0])

    return urls[1:]  # Skip the header row
